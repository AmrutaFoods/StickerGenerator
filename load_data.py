#!/usr/bin/env python3

'''data loading from excel'''

from openpyxl import load_workbook

def load_data():
    '''method to load data from excel'''
    try:
        # Load the Excel workbook
        workbook = load_workbook('items_list.xlsx')
        sheet = workbook.active

        items_english = []
        items = []
        weights = []
        mrps = []

        # Iterate over each row in the sheet
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Extract data from the row
            item_english = row[0]
            item = row[1]
            weight_str = row[2]
            mrp_str = row[3]


            # Check if weight and mrp strings are empty
            if not weight_str or not mrp_str:
                raise ValueError(f"Weight or MRP string is empty for item '{item}'.")

            # Split the comma-separated strings into lists and remove spaces from each item
            weight_list = [weight.strip() for weight in weight_str.split(',')]
            mrp_list = [mrp.strip() for mrp in mrp_str.split(',')]

            #     # Check if the number of net weights and MRPs match
            if len(weight_list) != len(mrp_list):
                raise ValueError(f"Number of net weights and MRPs are not same for Item - {item}.")

            # Append the data to the lists
            items_english.append(item_english)
            items.append(item)
            weights.append(weight_list)
            mrps.append(mrp_list)
        return items_english, items, weights, mrps
    except FileNotFoundError:
        print("Error: The specified Excel file was not found.")
    except ValueError as e:
        print(f"Error: {e}")
